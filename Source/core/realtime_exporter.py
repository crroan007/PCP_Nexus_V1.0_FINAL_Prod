import csv
import os
import json
import datetime
import shutil
import threading
from pathlib import Path
from core import app_paths
from core.secure_config import conf

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[EXPORT] Warning: openpyxl not available, Excel export disabled")

class RealtimeExporter:
    """
    Real-time CSV and Excel export engine.
    Appends job data to live export files as jobs are processed.
    
    CSV Format: Simplified 8-column format for import compatibility
    Excel Format: Full detailed columns for audit/review
    
    v4.3 Compliance Features:
    - Phase-specific network share output paths (configurable)
    - CSV rotation & archive lifecycle with configurable archive paths
    - Scheduled CSV placement (configurable per phase)
    - Auto-email CSV delivery (configurable per phase)
    """
    
    # CSV Header (simplified for import)
    CSV_HEADER = [
        "Envelope_Num", "Case_Num", "Date_Submitted", "Time_Submitted",
        "Date_Accepted", "Time_Accepted", "Lead_Document", "PCP_Job_Num"
    ]
    
    # Excel Header (full details)
    EXCEL_HEADER = [
        "ID", "Timestamp", "Filename", "Envelope_ID", "Case_Number", 
        "Lead_Doc", "Orig_Input", "Orig_Prefix", "New_Prefix", "Status", "Outcome", "Next_Steps",
        "Has_Comments", "Comments"
    ]
    
    # ──────────────────────────────────────────────
    # v4.3: Default values — all overridable via config.json
    # ──────────────────────────────────────────────
    DEFAULT_NETWORK_PATHS = {
        "Phase1": r"\\172.31.47.151\psaffidavits\AI_Inbound\File_Stamped_Affidavits",
        "Phase2": r"\\172.31.47.151\psaffidavits\AI_Inbound\File_Stamped_Petitions",
    }
    
    DEFAULT_ARCHIVE_PATHS = {
        "Phase1": r"\\172.31.47.151\psaffidavits\AI_Inbound\File_Stamped_Affidavits\Archive",
        "Phase2": r"\\172.31.47.151\psaffidavits\AI_Inbound\File_Stamped_Petitions\Archive",
    }
    
    # Fallback schedule — overridden by config keys schedule.phase1_* / schedule.phase2_*
    DEFAULT_SCHEDULE_CONFIG = {
        "Phase1": {"mode": "daily", "hour": 22, "minute": 0},   # 10:00 PM daily
        "Phase2": {"mode": "hourly"},                             # Every hour
    }
    
    def __init__(self):
        # Determine output directory from config or default to ProgramData
        # Config key: paths.realtime_exports (falls back to paths.output, then default)
        default_output = r"C:\ProgramData\PCP-Automation\Output"
        self.output_dir = conf.get("paths.realtime_exports") or conf.get("paths.output") or default_output
        if not self.output_dir or not self.output_dir.strip():
            self.output_dir = default_output
        os.makedirs(self.output_dir, exist_ok=True)
        
        # v4.3: Phase-specific network share paths (configurable)
        self.phase_output_paths = {
            "Phase1": conf.get("paths.csv_output_phase1") or self.DEFAULT_NETWORK_PATHS["Phase1"],
            "Phase2": conf.get("paths.csv_output_phase2") or self.DEFAULT_NETWORK_PATHS["Phase2"],
        }
        
        # v4.3: Phase-specific archive paths (configurable)
        self.phase_archive_paths = {
            "Phase1": conf.get("paths.csv_archive_phase1") or self.DEFAULT_ARCHIVE_PATHS["Phase1"],
            "Phase2": conf.get("paths.csv_archive_phase2") or self.DEFAULT_ARCHIVE_PATHS["Phase2"],
        }
        
        # v4.3: Schedule configuration (configurable per phase)
        self.schedule_config = {
            "Phase1": {
                "mode": conf.get("schedule.phase1_mode") or self.DEFAULT_SCHEDULE_CONFIG["Phase1"]["mode"],
                "hour": int(conf.get("schedule.phase1_hour") or self.DEFAULT_SCHEDULE_CONFIG["Phase1"]["hour"]),
                "minute": int(conf.get("schedule.phase1_minute") or self.DEFAULT_SCHEDULE_CONFIG["Phase1"]["minute"]),
                "interval_minutes": int(conf.get("schedule.phase1_interval_minutes") or 3),
            },
            "Phase2": {
                "mode": conf.get("schedule.phase2_mode") or self.DEFAULT_SCHEDULE_CONFIG["Phase2"]["mode"],
                "interval_minutes": int(conf.get("schedule.phase2_interval_minutes") or 3),
            },
        }
        
        # v4.3: Auto-email configuration (configurable per phase)
        self.auto_email_config = {
            "Phase1": {
                "enabled": conf.get("email.phase1_enabled") or False,
                "recipient": conf.get("email.phase1_recipient") or "",
            },
            "Phase2": {
                "enabled": conf.get("email.phase2_enabled") or False,
                "recipient": conf.get("email.phase2_recipient") or "",
            },
        }
        
        # v4.3: Scheduling state — tracks last placement time per phase
        self._last_placement = {
            "Phase1": None,
            "Phase2": None,
        }
        self._schedule_lock = threading.Lock()
        
        # Generate today's filename base
        today_str = datetime.datetime.now().strftime("%Y%m%d")
        
        # v4.3: Phase-specific CSV files
        self.csv_paths = {
            "Phase1": os.path.join(self.output_dir, f"pcp_phase1_{today_str}.csv"),
            "Phase2": os.path.join(self.output_dir, f"pcp_phase2_{today_str}.csv"),
        }
        # Legacy compat: keep csv_path pointing to Phase1
        self.csv_path = self.csv_paths["Phase1"]
        
        self.excel_filename = f"pcp_activity_{today_str}.xlsx"
        self.excel_path = os.path.join(self.output_dir, self.excel_filename)
        
        # Ensure both phase CSVs have headers
        self._ensure_csv_headers()
        
        # Ensure Excel workbook exists
        if EXCEL_AVAILABLE:
            self._ensure_excel_workbook()
        
        print(f"[EXPORT] Config loaded — Phase1 share: {self.phase_output_paths['Phase1']}")
        print(f"[EXPORT] Config loaded — Phase2 share: {self.phase_output_paths['Phase2']}")
        print(f"[EXPORT] Config loaded — Phase1 schedule: {self.schedule_config['Phase1']}")
        print(f"[EXPORT] Config loaded — Phase2 schedule: {self.schedule_config['Phase2']}")
    
    def _ensure_csv_headers(self):
        """Creates phase-specific CSVs with header if they don't exist."""
        for phase, path in self.csv_paths.items():
            if not os.path.exists(path):
                with open(path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.CSV_HEADER)
    
    # Legacy alias
    def _ensure_csv_header(self):
        self._ensure_csv_headers()
    
    def _ensure_excel_workbook(self):
        """Creates Excel workbook with full header if it doesn't exist."""
        if not os.path.exists(self.excel_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "PCP Activity"
            
            # Header
            ws.append(self.EXCEL_HEADER)
            
            # Style header row
            header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            wb.save(self.excel_path)
    
    # ──────────────────────────────────────────────
    # v4.3: CSV Rotation & Archive Lifecycle
    # ──────────────────────────────────────────────
    
    def rotate_csv_to_share(self, phase, source_csv_path=None):
        """
        Rotate-and-replace CSV placement to network share (v4.3 compliance).
        
        Steps:
        1. Check if target network share path is accessible
        2. If an existing CSV exists at the target, move it to an Archive subfolder
        
        Uses phase-specific source CSV if source_csv_path not provided.
        3. Copy the fresh CSV to the target location
        
        Args:
            phase: "Phase1" or "Phase2"
            source_csv_path: Path to the source CSV (defaults to self.csv_path)
            
        Returns:
            (success: bool, message: str)
        """
        target_dir = self.phase_output_paths.get(phase)
        source = source_csv_path or self.csv_paths.get(phase, self.csv_path)
        
        if not os.path.exists(source):
            return False, f"Source CSV not found: {source}"
        
        target_dir = self.phase_output_paths.get(phase)
        if not target_dir:
            return False, f"No output path configured for {phase}"
        
        try:
            # Ensure target directory exists (may fail if network share is offline)
            os.makedirs(target_dir, exist_ok=True)
            
            target_csv = os.path.join(target_dir, os.path.basename(source))
            
            # Step 1: Archive existing CSV if present
            if os.path.exists(target_csv):
                archive_dir = self.phase_archive_paths.get(phase) or os.path.join(target_dir, "Archive")
                os.makedirs(archive_dir, exist_ok=True)
                
                # Timestamp the archived copy
                ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
                archive_name = f"{Path(target_csv).stem}_{ts}.csv"
                archive_path = os.path.join(archive_dir, archive_name)
                
                shutil.move(target_csv, archive_path)
                print(f"[EXPORT] Archived existing CSV → {archive_path}")
            
            # Step 2: MOVE local CSV to the share (not copy)
            shutil.move(source, target_csv)
            print(f"[EXPORT] Moved CSV → {target_csv}")
            
            # Step 3: Reset local CSV with blank + headers
            with open(source, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(self.CSV_HEADER)
            print(f"[EXPORT] Reset local CSV → {source}")
            
            return True, f"CSV placed at {target_csv}, local reset"
            
        except PermissionError as e:
            return False, f"Permission denied accessing {target_dir}: {e}"
        except OSError as e:
            return False, f"Network share unavailable ({target_dir}): {e}"
        except Exception as e:
            return False, f"CSV rotation error: {e}"
    
    # ──────────────────────────────────────────────
    # v4.3: Scheduled CSV Placement
    # ──────────────────────────────────────────────
    
    def check_and_place_scheduled(self, phase):
        """
        Check if it's time to place a CSV for the given phase,
        and rotate if the schedule condition is met.
        
        Supported modes:
          - daily: Place once per day at a specific hour/minute
          - hourly: Place every hour
          - minutes: Place every N minutes (configurable via schedule.phaseX_interval_minutes)
        
        Args:
            phase: "Phase1" or "Phase2"
            
        Returns:
            (placed: bool, message: str)
        """
        now = datetime.datetime.now()
        schedule = self.schedule_config.get(phase)
        
        if not schedule:
            return False, f"No schedule configured for {phase}"
        
        with self._schedule_lock:
            last = self._last_placement.get(phase)
            
            if schedule["mode"] == "daily":
                # Phase 1: Daily at specific hour
                target_hour = schedule.get("hour", 22)
                target_minute = schedule.get("minute", 0)
                
                # Check if we're past the target time and haven't placed today
                if now.hour >= target_hour and now.minute >= target_minute:
                    if last is None or last.date() < now.date():
                        # Time to place
                        success, msg = self.rotate_csv_to_share(phase)
                        if success:
                            self._last_placement[phase] = now
                        return success, msg
                        
            elif schedule["mode"] == "hourly":
                # Phase 2: Every hour
                if last is None or (now - last).total_seconds() >= 3600:
                    success, msg = self.rotate_csv_to_share(phase)
                    if success:
                        self._last_placement[phase] = now
                    return success, msg
            
            elif schedule["mode"] == "minutes":
                # v4.3.3: Every N minutes (configurable, for testing or aggressive placement)
                interval = schedule.get("interval_minutes", 3)
                if last is None or (now - last).total_seconds() >= (interval * 60):
                    success, msg = self.rotate_csv_to_share(phase)
                    if success:
                        self._last_placement[phase] = now
                    return success, msg
            
            return False, "Not yet time for scheduled placement"
    
    def run_scheduled_placements(self):
        """
        Run all scheduled CSV placements. Called by the engine's reporter loop.
        Returns a dict of results per phase.
        """
        results = {}
        for phase in ["Phase1", "Phase2"]:
            placed, msg = self.check_and_place_scheduled(phase)
            results[phase] = {"placed": placed, "message": msg}
            if placed:
                print(f"[EXPORT] Scheduled placement for {phase}: {msg}")
        return results
    
    def _parse_datetime(self, raw_datetime_str):
        """
        Parse raw datetime string into separate date and time strings.
        Returns (date_str, time_str) tuple.
        """
        if not raw_datetime_str:
            return ("", "")
        
        try:
            # Try common formats
            for fmt in ["%Y-%m-%d %H:%M:%S", "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %H:%M:%S", 
                       "%Y-%m-%dT%H:%M:%S", "%m/%d/%y %H:%M:%S", "%m/%d/%y %I:%M:%S %p"]:
                try:
                    dt = datetime.datetime.strptime(raw_datetime_str.split('.')[0], fmt)
                    return (dt.strftime("%m/%d/%Y"), dt.strftime("%H:%M:%S"))
                except ValueError:
                    continue
            
            # Fallback: try to split by space
            parts = raw_datetime_str.strip().split(' ')
            if len(parts) >= 2:
                return (parts[0], ' '.join(parts[1:]))
            return (raw_datetime_str, "")
        except:
            return (raw_datetime_str, "")
    
    def export_job(self, job_id, filename, status, metadata_json, service_type='AFFIDAVITS'):
        """
        Appends a single job record to both CSV and Excel in real-time.
        
        CSV: Phase-specific files — Phase 1 on ARCHIVED, Phase 2 on FILED
        Excel: Full detailed format (14 columns) - All statuses
        """
        try:
            # Parse metadata
            try:
                metadata = json.loads(metadata_json) if metadata_json else {}
            except:
                metadata = {}
            
            # === SHARED FIELD EXTRACTION ===
            # Check multiple possible keys for envelope (envelope, envelope_num, envelope_id)
            envelope_num = metadata.get('envelope', metadata.get('envelope_num', metadata.get('envelope_id', '-')))

            case_num = metadata.get('case_num', '-')
            lead_doc = metadata.get('lead_doc', metadata.get('lead_document', metadata.get('subject', '-')))
            pcp_job_num = metadata.get('pcp_job_num', metadata.get('job_num', '-'))
            
            # Parse submitted/accepted dates
            date_submitted_raw = metadata.get('date_submitted_raw', '')
            date_accepted_raw = metadata.get('date_accepted_raw', '')
            
            date_submitted, time_submitted = self._parse_datetime(date_submitted_raw)
            date_accepted, time_accepted = self._parse_datetime(date_accepted_raw)
            
            # === CSV ROW (SIMPLIFIED) ===
            csv_row = [
                str(envelope_num),
                str(case_num),
                date_submitted,
                time_submitted,
                date_accepted,
                time_accepted,
                str(lead_doc),
                str(pcp_job_num)
            ]
            
            # === PHASE-SPECIFIC CSV WRITE ===
            # Phase 1 (AFFIDAVITS): Write on ARCHIVED
            # Phase 2 (PROJECT_2): Write on FILED
            phase = "Phase1" if service_type == 'AFFIDAVITS' else "Phase2"
            csv_trigger_status = "ARCHIVED" if phase == "Phase1" else "FILED"
            
            if status == csv_trigger_status:
                csv_path = self.csv_paths.get(phase, self.csv_path)
                with open(csv_path, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(csv_row)

            
            # === EXCEL ROW (FULL DETAILS) ===
            if EXCEL_AVAILABLE:
                import re
                
                orig_input = metadata.get('original_filename', filename)
                if orig_input and '\\' in orig_input:
                    orig_input = os.path.basename(orig_input)
                
                # Prefixes - Extract from Lead Document
                # Orig_Prefix = first 4 chars of Lead_Doc (Column F, e.g., AX02 from AX02A26101444.PDF)
                lead_doc_str = str(lead_doc) if lead_doc else ""
                orig_prefix = lead_doc_str[:4].upper() if len(lead_doc_str) >= 4 else "-"
                    
                # Destination prefix: first 4 chars of final filename (for comparison)
                dest_prefix = filename[:4].upper() if filename and len(filename) >= 4 else "-"
                
                # Status mapping
                outcome_map = {
                    "ARCHIVED": "Filing Accepted / Verified",
                    "VERIFIED": "Verification Successful",
                    "QA_FAILED": "Quality Check Failed",
                    "EXCEPTION": "Processing Error",
                    "NEW": "Pending Processing",
                    "FILED": "Filing Complete"
                }
                outcome = outcome_map.get(status, status)
                
                next_steps_map = {
                    "ARCHIVED": "Archived Successfully",
                    "VERIFIED": "Ready for Archive",
                    "QA_FAILED": "Manual Review Required",
                    "EXCEPTION": "Error Resolution Needed",
                    "NEW": "In Queue",
                    "FILED": "Complete"
                }
                next_steps = next_steps_map.get(status, "Processing")
                
                # Comments
                raw_comments = metadata.get('raw_comments', metadata.get('comments', ''))
                has_comments = "Yes" if raw_comments and raw_comments.strip() and raw_comments.strip() != '-' else "No"
                comment_text = raw_comments.strip() if raw_comments else "-"
                
                timestamp = datetime.datetime.now().strftime("%H:%M")
                excel_row = [
                    str(job_id),
                    timestamp,
                    filename,
                    envelope_num,
                    case_num,
                    lead_doc,
                    orig_input,
                    orig_prefix,
                    dest_prefix,
                    status,
                    outcome,
                    next_steps,
                    has_comments,
                    comment_text
                ]
                
                try:
                    wb = load_workbook(self.excel_path)
                    ws = wb.active
                    ws.append(excel_row)
                    
                    # Auto-adjust column widths
                    for col in ws.columns:
                        max_length = 0
                        column = col[0].column_letter
                        for cell in col:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column].width = adjusted_width
                    
                    wb.save(self.excel_path)
                except Exception as e:
                    print(f"[EXPORT] Excel write error for job {job_id}: {e}")
            
            print(f"[EXPORT] Logged job #{job_id} ({phase}) to CSV and {self.excel_filename}")
            
        except Exception as e:
            print(f"[EXPORT] Error exporting job {job_id}: {e}")

# Global singleton instance
exporter = RealtimeExporter()
