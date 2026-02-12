import openpyxl

wb = openpyxl.load_workbook(r"D:\Homebrew Apps\PCP V1.0 (final) - Copy\test results.xlsx", data_only=True)
ws = wb.active

headers = []
for col in range(1, ws.max_column + 1):
    val = ws.cell(row=1, column=col).value
    headers.append(str(val).strip() if val else f"Col{col}")

lines = []
lines.append(f"Sheet: {ws.title} | {ws.max_row-1} data rows | {ws.max_column} cols")
lines.append(f"Headers: {headers}")
lines.append("=" * 150)

for row in range(2, ws.max_row + 1):
    vals = []
    for col in range(1, ws.max_column + 1):
        v = ws.cell(row=col, column=col).value if False else ws.cell(row=row, column=col).value
        if v is None:
            vals.append("")
        elif isinstance(v, float):
            vals.append(str(int(v)) if v == int(v) else str(v))
        else:
            vals.append(str(v).strip())
    
    if not any(vals):
        continue
    
    envelope = vals[0] if len(vals) > 0 else ""
    output = vals[1] if len(vals) > 1 else ""
    job_name = vals[2] if len(vals) > 2 else ""
    pages = vals[3] if len(vals) > 3 else ""
    sources = vals[4] if len(vals) > 4 else ""
    src_pages = vals[5] if len(vals) > 5 else ""
    check = vals[6] if len(vals) > 6 else ""
    status = vals[7] if len(vals) > 7 else ""
    source_docs = vals[8] if len(vals) > 8 else ""
    
    lines.append(f"R{row:3d} | ENV={envelope:12s} | OUT={output:30s} | JOB={job_name:20s} | PG={pages:4s} | SRC={sources:2s} | SRCPG={src_pages:4s} | CHK={check:2s} | {status:6s} | {source_docs}")

lines.append("=" * 150)
lines.append(f"Total: {ws.max_row - 1}")

out_path = r"D:\Homebrew Apps\PCP V1.0 (final) - Copy\PCP_USB_Backup_v1.0\Source\_results_clean.txt"
with open(out_path, 'w', encoding='ascii', errors='replace') as f:
    f.write('\n'.join(lines))

print(f"Written to {out_path}")
